import time
from selenium import webdriver
from bs4 import BeautifulSoup
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import openpyxl as op
from openpyxl import Workbook


#가오몬 1060 PRO
chrome_options = Options()# 크롬 브라우저에 적용할 옵션들을 위한 옵션 선언
driver = webdriver.Chrome('chromedriver.exe')# 크롬브라우저 띄우기
driver.maximize_window()# 창크기 최대화

URL = 'https://www.gmarket.co.kr/'# 크롬브라우저 띄울 때 시작화면 설정 주소


driver.get(URL) #get함수를 사용해 해당 URL을 브라우저에 띄운다

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
# 입력필드를 선택한다.
search_word = '가오몬 1060PRO'
search_box.send_keys(search_word)
# 입력필드에 sendKey를 사용해 search_word에 저장되어있는 검색어를 입력한다.

driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()
# 검색버튼을 선택한다.

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser') # URL에 해당하는 페이지의 HTML을 가져온다.
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')
#list변수에 html정보를 가져올 상위 태그를 select함수를 사용하여 저장한다.


wb = op.Workbook()# 새로운 workbook 객체 생성
ws = wb.active# 활성화 된 시트를 워크시트 객체로 생성
ws.title = "G마켓"# 시트명 지정
ws.append(['브랜드명','상품명','구매수량','판매처','상품금액'])
# append함수를 사용해 엑셀 첫번째 컬럼에 순차적으로 각각의 타이틀명을 추가해준다.

  
for b in list2:    
  for v in list: #이중 for문을 사용해 화면 첫단 부터 끝단까지의 정보까지 크롤링하는 작업을 반복한다.       
    item_word = b.select("button")[0]['data-montelena-keyword']
    # 검색한 브랜드명 값을 가져온다.
    item_name = v.find('span', {"class": "text__item"}).text       
    # 상품명을 가져온다.

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue
    # 구매건수를 가져오는데, 구매건수가 없는건 제외하고 구매건수만 있는 정보만 추출하여 계속 진행한다.

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
    # 판매처를 가져오는데, '미니샵으로 이동합니다'는 공백으로 하여 값을 없앤다.
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 
    # 판매상품의 판매가격을 가져온다.

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    # 가져온 정보를 출력
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    # 현재 활성화된 엑셀 워크시트에 각각의 정보들을 append함수를 사용해 추가해준다.
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
    # 엑셀파일 절대경로에 저장.
    
#가오몬 S620
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '가오몬 S620'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
# load_workbook = 새로운 엑셀파일에 저장하는 것이 아닌, 기존 엑셀파일에 저장한다.
ws = wb.active

for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#가오몬 M10K
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '가오몬 M10K'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")


#가오몬 M6
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '가오몬 M6'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#가오몬 WH580
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '가오몬 WH580'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")


#vinsa
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = 'vinsa'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#보스토 T1060
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '보스토 T1060'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#베이크 A15
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '베이크 A15'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#베이크 A30
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '베이크 A30'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#베이크 A50
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '베이크 A50'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#베이크 VK640
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '베이크 VK640'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#베이크 VK1060
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '베이크 VK1060'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#베이크 S640
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '베이크 S640'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#이지드로잉 1060plus
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '이지드로잉 1060plus'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#이지드로잉 노트
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '이지드로잉 노트'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()


pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')

wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active

for b in list2:    
  for v in list:          
       item_word = b.select("button")[0]['data-montelena-keyword']

       item_name = v.find('span', {"class": "text__item"}).text       

       try:   
          item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
       except Exception as e:
          continue       
      
       item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
       new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
       item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

       print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
       ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
       wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#이지드로잉 지니
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '이지드로잉 지니'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#XPPEN DECO01
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = 'XPPEN DECO01'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#XPPEN DECO02
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = 'XPPEN DECO02'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#XPPEN Deco Pro S
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = 'XPPEN Deco Pro S'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']    
    new_item_word = item_word.replace('Deco Pro S', 'DecoProS')

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(new_item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([new_item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#XPPEN Deco Pro M
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = 'XPPEN Deco Pro M'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']
    new_item_word = item_word.replace('Deco Pro M', 'DecoProM') 

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(new_item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([new_item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#휴이온 RTS300
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '휴이온 RTS300'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#휴이온 RTM500
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '휴이온 RTM500'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#휴이온 RTP700
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '휴이온 RTP700'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#휴이온 H640P
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '휴이온 H640P'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#휴이온 H950P
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '휴이온 H950P'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
