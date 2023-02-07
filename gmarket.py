import time
from selenium import webdriver
from bs4 import BeautifulSoup
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import openpyxl as op
from openpyxl import Workbook
import pandas as pd
import requests
from bs4 import BeautifulSoup

#가오몬 1060 PRO
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '가오몬 1060PRO'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.Workbook()
ws = wb.active
ws.title = "G마켓"
ws.append(['브랜드명','상품명','구매수량','판매처','상품금액'])

  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
    
#가오몬 S620
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '가오몬 S620'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active

for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#가오몬 M10K
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '가오몬 M10K'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")


#가오몬 M6
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '가오몬 M6'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#가오몬 WH580
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '가오몬 WH580'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")


#vinsa
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = 'vinsa'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#보스토 T1060
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '보스토 T1060'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#베이크 A15
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '베이크 A15'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#베이크 A30
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '베이크 A30'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#베이크 A50
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '베이크 A50'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#베이크 VK640
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '베이크 VK640'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#베이크 VK1060
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '베이크 VK1060'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#베이크 S640
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '베이크 S640'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#이지드로잉 1060plus
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '이지드로잉 1060plus'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#이지드로잉 노트
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '이지드로잉 노트'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()


pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')

wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active

for b in list2:    
  for v in list:          
       item_word = b.select("button")[0]['data-montelena-keyword']

       item_name = v.find('span', {"class": "text__item"}).text       

       try:   
          item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
       except Exception as e:
          continue       
      
       item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
       new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
       item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

       print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
       ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
       wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#이지드로잉 지니
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '이지드로잉 지니'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#XPPEN DECO01
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = 'XPPEN DECO01'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#XPPEN DECO02
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = 'XPPEN DECO02'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#XPPEN Deco Pro S
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = 'XPPEN Deco Pro S'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']    
    new_item_word = item_word.replace('Deco Pro S', 'DecoProS')

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(new_item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([new_item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#XPPEN Deco Pro M
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = 'XPPEN Deco Pro M'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']
    new_item_word = item_word.replace('Deco Pro M', 'DecoProM') 

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(new_item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([new_item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#휴이온 RTS300
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '휴이온 RTS300'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#휴이온 RTM500
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '휴이온 RTM500'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#휴이온 RTP700
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '휴이온 RTP700'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#휴이온 H640P
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '휴이온 H640P'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")

#휴이온 H950P
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://www.gmarket.co.kr/'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/input')
search_word = '휴이온 H950P'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="skip-navigation-search"]/span/button').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')
list = soup.select('div.box__information')
list2 = soup.select('span.box__search-input')


wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
ws = wb.active
  
for b in list2:    
  for v in list:        
    item_word = b.select("button")[0]['data-montelena-keyword']

    item_name = v.find('span', {"class": "text__item"}).text       

    try:   
        item_pq = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information-score > ul > li.list-item.list-item__pay-count > span.text').text
    except Exception as e:
        continue

    item_seller = v.select_one('#section__inner-content-body-container > div > div > div.box__item-container > div.box__information > div.box__information_seller > a')['title'] 
    new_item_seller = item_seller.replace(' 미니샵으로 이동합니다', '')
      
    item_price = v.select_one('#section__inner-content-body-container > div > div > div > div.box__information > div.box__information-major > div.box__item-price > div > strong').text 

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_gmarket.xlsx")
