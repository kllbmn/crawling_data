import time
from selenium import webdriver
from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import csv
from PIL import ImageGrab, Image
import os
import openpyxl as op
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as pi



#CTL-472
chrome_options = Options() # 크롬 브라우저에 적용할 옵션들을 위한 옵션 선언

driver = webdriver.Chrome('chromedriver.exe') # 크롬브라우저 띄우기
driver.maximize_window() # 창크기 최대화

URL = 'https://shopping.naver.com/home' # 크롬브라우저 띄울 때 시작화면 설정 주소
driver.get(URL) # get함수를 사용해 해당 URL을 브라우저에 띄운다.

time.sleep(2) # 2초 대기

driver.find_element(By.CLASS_NAME,'_searchInput_search_text_fSuJ6').send_keys("CTL-472")
# 입력필드를 선택하고 sendkey로 검색어를 입력한다.

# find_element는 조건에 일치하는 가장 첫번째 요소를 반환
# find_elements는 조건에 일치하는 모든 요소를 list형태로 반환

driver.find_element(By.CLASS_NAME,'_searchInput_button_search_h79Dk').click()
# 검색버튼을 클릭한다.

driver.find_element(By.XPATH,'//*[@id="__next"]/div/div[2]/div/div[3]/div[1]/ul/div/div[1]/li/div/div[2]/div[2]/strong/a').click()
#최저가 버튼을 클릭한다.

time.sleep(2)

img = ImageGrab.grab() # 이미지 캡처, 전체화면을 캡처하기위해 ()안에 좌표값 없음
img.save(r"C:\Users\user\Desktop\leehotddu\img\CTL-472.png") # 절대경로 방식으로 이미지 저장경로 설정
# r 표시는 유니코드 에러 발생 시 사용. 파이썬에서는 역 슬래시를 인식하지 못해 디코더 에러 발생
driver.quit() # 현재 띄워져 있는 크롬 탭 종료. quit()는 전체 탭 종료 close()는 현재 보고 있는 크롬 탭만 종료

wb = op.Workbook() # 새로운 workbook 객체 생성
ws = wb.active # 활성화 된 시트를 워크시트 객체로 생성

path = r"C:\Users\user\Desktop\leehotddu\img"
number_file = "CTL-472.png"

img1 = Image(path + "/" + number_file) # Image 클래스의 객체 img 선언

ws.add_image(img1, "A1") # 엑셀 A1셀에 이미지 추가
ws.title = "CTL-472" # 시트명 지정

wb.save(r"C:\Users\user\Desktop\leehotddu\lee.xlsx")

wb.close() # 파일 닫기
driver.quit()

time.sleep(1)

#CTL-672
chrome_options = Options()

driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://shopping.naver.com/home'
driver.get(URL)

driver.find_element(By.CLASS_NAME,'_searchInput_search_text_fSuJ6').send_keys("CTL-672")

driver.find_element(By.CLASS_NAME,'_searchInput_button_search_h79Dk').click()

driver.find_element(By.XPATH,'//*[@id="__next"]/div/div[2]/div/div[3]/div[1]/ul/div/div[1]/li/div/div[2]/div[2]/strong/a').click()

time.sleep(2)

img = ImageGrab.grab()
img.save(r"C:\Users\user\Desktop\leehotddu\img\CTL-672.png")

wb = op.load_workbook(r"C:\Users\user\Desktop\leehotddu\lee.xlsx")
ws = wb.create_sheet("CTL-672")
ws = wb.active

path = r"C:\Users\user\Desktop\leehotddu\img"
number_file = "CTL-672.png"

img1 = Image(path + "/" + number_file)

ws = wb['CTL-672']
ws.add_image(img1, "A1")

wb.save(r"C:\Users\user\Desktop\leehotddu\lee.xlsx")

wb.close()
driver.quit()

time.sleep(1)

#CTL-4100
chrome_options = Options()

driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://shopping.naver.com/home'
driver.get(URL)

driver.find_element(By.CLASS_NAME,'_searchInput_search_text_fSuJ6').send_keys("CTL-4100")

driver.find_element(By.CLASS_NAME,'_searchInput_button_search_h79Dk').click()

driver.find_element(By.XPATH,'//*[@id="__next"]/div/div[2]/div/div[3]/div[1]/ul/div/div[1]/li/div/div[2]/div[2]/strong/a').click()

time.sleep(2)

img = ImageGrab.grab()
img.save(r"C:\Users\user\Desktop\leehotddu\img\CTL-4100.png")

wb = op.load_workbook(r"C:\Users\user\Desktop\leehotddu\lee.xlsx")
ws = wb.create_sheet("CTL-4100")
ws = wb.active

path = r"C:\Users\user\Desktop\leehotddu\img"
number_file = "CTL-4100.png"

img1 = Image(path + "/" + number_file)

ws = wb['CTL-4100']
ws.add_image(img1, "A1")

wb.save(r"C:\Users\user\Desktop\leehotddu\lee.xlsx")

wb.close()
driver.quit()

time.sleep(1)

#CTL-6100
chrome_options = Options()

driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://shopping.naver.com/home'
driver.get(URL)

driver.find_element(By.CLASS_NAME,'_searchInput_search_text_fSuJ6').send_keys("CTL-6100")

driver.find_element(By.CLASS_NAME,'_searchInput_button_search_h79Dk').click()

driver.find_element(By.XPATH,'//*[@id="__next"]/div/div[2]/div/div[3]/div[1]/ul/div/div[1]/li/div/div[2]/div[2]/strong/a').click()

time.sleep(2)

img = ImageGrab.grab()
img.save(r"C:\Users\user\Desktop\leehotddu\img\CTL-6100.png")

wb = op.load_workbook(r"C:\Users\user\Desktop\leehotddu\lee.xlsx")
ws = wb.create_sheet("CTL-6100")
ws = wb.active

path = r"C:\Users\user\Desktop\leehotddu\img"
number_file = "CTL-6100.png"

img1 = Image(path + "/" + number_file)

ws = wb['CTL-6100']
ws.add_image(img1, "A1")

wb.save(r"C:\Users\user\Desktop\leehotddu\lee.xlsx")

wb.close()
driver.quit()

time.sleep(1)

#CTL-4100WL
chrome_options = Options()

driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://shopping.naver.com/home'
driver.get(URL)

driver.find_element(By.CLASS_NAME,'_searchInput_search_text_fSuJ6').send_keys("CTL-4100WL")

driver.find_element(By.CLASS_NAME,'_searchInput_button_search_h79Dk').click()

driver.find_element(By.XPATH,'//*[@id="__next"]/div/div[2]/div/div[3]/div[1]/ul/div/div[1]/li/div/div[2]/div[2]/strong/a').click()

time.sleep(2)

img = ImageGrab.grab()
img.save(r"C:\Users\user\Desktop\leehotddu\img\CTL-4100WL.png")

wb = op.load_workbook(r"C:\Users\user\Desktop\leehotddu\lee.xlsx")
ws = wb.create_sheet("CTL-4100WL")
ws = wb.active

path = r"C:\Users\user\Desktop\leehotddu\img"
number_file = "CTL-4100WL.png"

img1 = Image(path + "/" + number_file)

ws = wb['CTL-4100WL']
ws.add_image(img1, "A1")

wb.save(r"C:\Users\user\Desktop\leehotddu\lee.xlsx")

wb.close()
driver.quit()

time.sleep(1)

#CTL-6100WL
chrome_options = Options()

driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://shopping.naver.com/home'
driver.get(URL)

driver.find_element(By.CLASS_NAME,'_searchInput_search_text_fSuJ6').send_keys("CTL-6100WL")

driver.find_element(By.CLASS_NAME,'_searchInput_button_search_h79Dk').click()

driver.find_element(By.XPATH,'//*[@id="__next"]/div/div[2]/div/div[3]/div[1]/ul/div/div[1]/li/div/div[2]/div[2]/strong/a').click()

time.sleep(2)

img = ImageGrab.grab()
img.save(r"C:\Users\user\Desktop\leehotddu\img\CTL-6100WL.png")

wb = op.load_workbook(r"C:\Users\user\Desktop\leehotddu\lee.xlsx")
ws = wb.create_sheet("CTL-6100WL")
ws = wb.active

path = r"C:\Users\user\Desktop\leehotddu\img"
number_file = "CTL-6100WL.png"

img1 = Image(path + "/" + number_file)

ws = wb['CTL-6100WL']
ws.add_image(img1, "A1")

wb.save(r"C:\Users\user\Desktop\leehotddu\lee.xlsx")

wb.close()
driver.quit()

time.sleep(1)

##DTC-133
chrome_options = Options()

driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://shopping.naver.com/home'
driver.get(URL)

driver.find_element(By.CLASS_NAME,'_searchInput_search_text_fSuJ6').send_keys("DTC-133")

driver.find_element(By.CLASS_NAME,'_searchInput_button_search_h79Dk').click()

driver.find_element(By.XPATH,'//*[@id="__next"]/div/div[2]/div/div[3]/div[1]/ul/div/div[1]/li/div/div[2]/div[2]/strong/a').click()

time.sleep(2)

img = ImageGrab.grab()
img.save(r"C:\Users\user\Desktop\leehotddu\img\DTC-133.png")

wb = op.load_workbook(r"C:\Users\user\Desktop\leehotddu\lee.xlsx")
ws = wb.create_sheet("DTC-133")
ws = wb.active

path = r"C:\Users\user\Desktop\leehotddu\img"
number_file = "DTC-133.png"

img1 = Image(path + "/" + number_file)

ws = wb['DTC-133']
ws.add_image(img1, "A1")

wb.save(r"C:\Users\user\Desktop\leehotddu\lee.xlsx")

wb.close()
driver.quit()