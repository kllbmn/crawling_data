import time
from selenium import webdriver
from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import csv
from PIL import ImageGrab, Image
import os
import openpyxl as op
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as pi



#CTL-472
chrome_options = Options()

driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://shopping.naver.com/home'
driver.get(URL)

time.sleep(2)

driver.find_element(By.CLASS_NAME,'_searchInput_search_text_fSuJ6').send_keys("CTL-472")

driver.find_element(By.CLASS_NAME,'_searchInput_button_search_h79Dk').click()

driver.find_element(By.XPATH,'//*[@id="__next"]/div/div[2]/div/div[3]/div[1]/ul/div/div[1]/li/div/div[2]/div[2]/strong/a').click()

time.sleep(2)

img = ImageGrab.grab()
img.save(r"C:\Users\hh_jung\Desktop\leehotddu\img\CTL-472.png")
driver.quit()

wb = op.Workbook()
ws = wb.active

path = r"C:\Users\hh_jung\Desktop\leehotddu\img"
number_file = "CTL-472.png"

img1 = Image(path + "/" + number_file)

ws.add_image(img1, "A1")
ws.title = "CTL-472"

wb.save(r"C:\Users\hh_jung\Desktop\leehotddu\lee.xlsx")

wb.close()
driver.quit()

time.sleep(1)

#CTL-672
chrome_options = Options()

driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://shopping.naver.com/home'
driver.get(URL)

driver.find_element(By.CLASS_NAME,'_searchInput_search_text_fSuJ6').send_keys("CTL-672")

driver.find_element(By.CLASS_NAME,'_searchInput_button_search_h79Dk').click()

driver.find_element(By.XPATH,'//*[@id="__next"]/div/div[2]/div/div[3]/div[1]/ul/div/div[1]/li/div/div[2]/div[2]/strong/a').click()

time.sleep(2)

img = ImageGrab.grab()
img.save(r"C:\Users\hh_jung\Desktop\leehotddu\img\CTL-672.png")

wb = op.load_workbook(r"C:\Users\hh_jung\Desktop\leehotddu\lee.xlsx")
ws = wb.create_sheet("CTL-672")
ws = wb.active

path = r"C:\Users\hh_jung\Desktop\leehotddu\img"
number_file = "CTL-672.png"

img1 = Image(path + "/" + number_file)

ws = wb['CTL-672']
ws.add_image(img1, "A1")

wb.save(r"C:\Users\hh_jung\Desktop\leehotddu\lee.xlsx")

wb.close()
driver.quit()

time.sleep(1)

#CTL-4100
chrome_options = Options()

driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://shopping.naver.com/home'
driver.get(URL)

driver.find_element(By.CLASS_NAME,'_searchInput_search_text_fSuJ6').send_keys("CTL-4100")

driver.find_element(By.CLASS_NAME,'_searchInput_button_search_h79Dk').click()

driver.find_element(By.XPATH,'//*[@id="__next"]/div/div[2]/div/div[3]/div[1]/ul/div/div[1]/li/div/div[2]/div[2]/strong/a').click()

time.sleep(2)

img = ImageGrab.grab()
img.save(r"C:\Users\hh_jung\Desktop\leehotddu\img\CTL-4100.png")

wb = op.load_workbook(r"C:\Users\hh_jung\Desktop\leehotddu\lee.xlsx")
ws = wb.create_sheet("CTL-4100")
ws = wb.active

path = r"C:\Users\hh_jung\Desktop\leehotddu\img"
number_file = "CTL-4100.png"

img1 = Image(path + "/" + number_file)

ws = wb['CTL-4100']
ws.add_image(img1, "A1")

wb.save(r"C:\Users\hh_jung\Desktop\leehotddu\lee.xlsx")

wb.close()
driver.quit()

time.sleep(1)

#CTL-6100
chrome_options = Options()

driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://shopping.naver.com/home'
driver.get(URL)

driver.find_element(By.CLASS_NAME,'_searchInput_search_text_fSuJ6').send_keys("CTL-6100")

driver.find_element(By.CLASS_NAME,'_searchInput_button_search_h79Dk').click()

driver.find_element(By.XPATH,'//*[@id="__next"]/div/div[2]/div/div[3]/div[1]/ul/div/div[1]/li/div/div[2]/div[2]/strong/a').click()

time.sleep(2)

img = ImageGrab.grab()
img.save(r"C:\Users\hh_jung\Desktop\leehotddu\img\CTL-6100.png")

wb = op.load_workbook(r"C:\Users\hh_jung\Desktop\leehotddu\lee.xlsx")
ws = wb.create_sheet("CTL-6100")
ws = wb.active

path = r"C:\Users\hh_jung\Desktop\leehotddu\img"
number_file = "CTL-6100.png"

img1 = Image(path + "/" + number_file)

ws = wb['CTL-6100']
ws.add_image(img1, "A1")

wb.save(r"C:\Users\hh_jung\Desktop\leehotddu\lee.xlsx")

wb.close()
driver.quit()

time.sleep(1)

#CTL-4100WL
chrome_options = Options()

driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://shopping.naver.com/home'
driver.get(URL)

driver.find_element(By.CLASS_NAME,'_searchInput_search_text_fSuJ6').send_keys("CTL-4100WL")

driver.find_element(By.CLASS_NAME,'_searchInput_button_search_h79Dk').click()

driver.find_element(By.XPATH,'//*[@id="__next"]/div/div[2]/div/div[3]/div[1]/ul/div/div[1]/li/div/div[2]/div[2]/strong/a').click()

time.sleep(2)

img = ImageGrab.grab()
img.save(r"C:\Users\hh_jung\Desktop\leehotddu\img\CTL-4100WL.png")

wb = op.load_workbook(r"C:\Users\hh_jung\Desktop\leehotddu\lee.xlsx")
ws = wb.create_sheet("CTL-4100WL")
ws = wb.active

path = r"C:\Users\hh_jung\Desktop\leehotddu\img"
number_file = "CTL-4100WL.png"

img1 = Image(path + "/" + number_file)

ws = wb['CTL-4100WL']
ws.add_image(img1, "A1")

wb.save(r"C:\Users\hh_jung\Desktop\leehotddu\lee.xlsx")

wb.close()
driver.quit()

time.sleep(1)

#CTL-6100WL
chrome_options = Options()

driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://shopping.naver.com/home'
driver.get(URL)

driver.find_element(By.CLASS_NAME,'_searchInput_search_text_fSuJ6').send_keys("CTL-6100WL")

driver.find_element(By.CLASS_NAME,'_searchInput_button_search_h79Dk').click()

driver.find_element(By.XPATH,'//*[@id="__next"]/div/div[2]/div/div[3]/div[1]/ul/div/div[1]/li/div/div[2]/div[2]/strong/a').click()

time.sleep(2)

img = ImageGrab.grab()
img.save(r"C:\Users\hh_jung\Desktop\leehotddu\img\CTL-6100WL.png")

wb = op.load_workbook(r"C:\Users\hh_jung\Desktop\leehotddu\lee.xlsx")
ws = wb.create_sheet("CTL-6100WL")
ws = wb.active

path = r"C:\Users\hh_jung\Desktop\leehotddu\img"
number_file = "CTL-6100WL.png"

img1 = Image(path + "/" + number_file)

ws = wb['CTL-6100WL']
ws.add_image(img1, "A1")

wb.save(r"C:\Users\hh_jung\Desktop\leehotddu\lee.xlsx")

wb.close()
driver.quit()

time.sleep(1)

#DTC-133
chrome_options = Options()

driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'https://shopping.naver.com/home'
driver.get(URL)

driver.find_element(By.CLASS_NAME,'_searchInput_search_text_fSuJ6').send_keys("DTC-133")

driver.find_element(By.CLASS_NAME,'_searchInput_button_search_h79Dk').click()

driver.find_element(By.XPATH,'//*[@id="__next"]/div/div[2]/div/div[3]/div[1]/ul/div/div[1]/li/div/div[2]/div[2]/strong/a').click()

time.sleep(2)

img = ImageGrab.grab()
img.save(r"C:\Users\hh_jung\Desktop\leehotddu\img\DTC-133.png")

wb = op.load_workbook(r"C:\Users\hh_jung\Desktop\leehotddu\lee.xlsx")
ws = wb.create_sheet("DTC-133")
ws = wb.active

path = r"C:\Users\hh_jung\Desktop\leehotddu\img"
number_file = "DTC-133.png"

img1 = Image(path + "/" + number_file)

ws = wb['DTC-133']
ws.add_image(img1, "A1")

wb.save(r"C:\Users\hh_jung\Desktop\leehotddu\lee.xlsx")

wb.close()
driver.quit()