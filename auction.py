import time
from selenium import webdriver
from bs4 import BeautifulSoup
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import openpyxl as op
from openpyxl import Workbook
import pandas as pd
import requests
from bs4 import BeautifulSoup



#가오몬 1060PRO
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'http://www.auction.co.kr/??pid=867&redirect=1'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="txtKeyword"]')
search_word = '가오몬 1060PRO'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="core_header"]/div/div[1]/form/div[1]/input[2]').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')

list = soup.select('div.section--itemcard_info')

wb = op.Workbook()
ws = wb.active
ws.title = "옥션"
ws.append(['브랜드명','상품명','구매수량','판매처','상품금액'])

     
for v in list:        
    item_word = soup.head.find("meta", {"name" : "keywords"}).get('content')

    item_name = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_major > div.area--itemcard_title > span > a > span.text--title').text     

    try:   
        item_pq = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_score > ul > li > span.text--buycnt').text
    except Exception as e:
        continue
    
    try:
        item_seller = v.select_one('a.link--shop').get_text()  
        new_item_seller = item_seller.replace('판매자', '')
    except Exception as e:
        new_item_seller = v.select_one('#section--inner_content_body_container > div > div.component.component--item_card.type--smiledelivery > div > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_smiledelivery > a > img')['alt']


    item_price = v.select_one('strong.text--price_seller').get_text()

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")
    

#가오몬 S620
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'http://www.auction.co.kr/??pid=867&redirect=1'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="txtKeyword"]')
search_word = '가오몬 S620'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="core_header"]/div/div[1]/form/div[1]/input[2]').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')

list = soup.select('div.section--itemcard_info')

wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")
ws = wb.active
     
for v in list:        
    item_word = soup.head.find("meta", {"name" : "keywords"}).get('content')

    item_name = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_major > div.area--itemcard_title > span > a > span.text--title').text     

    try:   
        item_pq = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_score > ul > li > span.text--buycnt').text
    except Exception as e:
        continue
    
    try:
        item_seller = v.select_one('a.link--shop').get_text()  
        new_item_seller = item_seller.replace('판매자', '')
    except Exception as e:
        new_item_seller = v.select_one('#section--inner_content_body_container > div > div.component.component--item_card.type--smiledelivery > div > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_smiledelivery > a > img')['alt']


    item_price = v.select_one('strong.text--price_seller').get_text()

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")


#가오몬 M10K
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'http://www.auction.co.kr/??pid=867&redirect=1'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="txtKeyword"]')
search_word = '가오몬 M10K'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="core_header"]/div/div[1]/form/div[1]/input[2]').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')

list = soup.select('div.section--itemcard_info')

wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")
ws = wb.active

     
for v in list:        
    item_word = soup.head.find("meta", {"name" : "keywords"}).get('content')

    item_name = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_major > div.area--itemcard_title > span > a > span.text--title').text     

    try:   
        item_pq = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_score > ul > li > span.text--buycnt').text
    except Exception as e:
        continue
    
    try:
        item_seller = v.select_one('a.link--shop').get_text()  
        new_item_seller = item_seller.replace('판매자', '')
    except Exception as e:
        new_item_seller = v.select_one('#section--inner_content_body_container > div > div.component.component--item_card.type--smiledelivery > div > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_smiledelivery > a > img')['alt']


    item_price = v.select_one('strong.text--price_seller').get_text()

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")


#가오몬 M6
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'http://www.auction.co.kr/??pid=867&redirect=1'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="txtKeyword"]')
search_word = '가오몬 M6'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="core_header"]/div/div[1]/form/div[1]/input[2]').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')

list = soup.select('div.section--itemcard_info')

wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")
ws = wb.active

     
for v in list:        
    item_word = soup.head.find("meta", {"name" : "keywords"}).get('content')

    item_name = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_major > div.area--itemcard_title > span > a > span.text--title').text     

    try:   
        item_pq = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_score > ul > li > span.text--buycnt').text
    except Exception as e:
        continue
    
    try:
        item_seller = v.select_one('a.link--shop').get_text()  
        new_item_seller = item_seller.replace('판매자', '')
    except Exception as e:
        new_item_seller = v.select_one('#section--inner_content_body_container > div > div.component.component--item_card.type--smiledelivery > div > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_smiledelivery > a > img')['alt']


    item_price = v.select_one('strong.text--price_seller').get_text()

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")


#가오몬 WH850
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'http://www.auction.co.kr/??pid=867&redirect=1'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="txtKeyword"]')
search_word = '가오몬 WH850'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="core_header"]/div/div[1]/form/div[1]/input[2]').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')

list = soup.select('div.section--itemcard_info')

wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")
ws = wb.active

     
for v in list:        
    item_word = soup.head.find("meta", {"name" : "keywords"}).get('content')

    item_name = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_major > div.area--itemcard_title > span > a > span.text--title').text     

    try:   
        item_pq = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_score > ul > li > span.text--buycnt').text
    except Exception as e:
        continue
    
    try:
        item_seller = v.select_one('a.link--shop').get_text()  
        new_item_seller = item_seller.replace('판매자', '')
    except Exception as e:
        new_item_seller = v.select_one('#section--inner_content_body_container > div > div.component.component--item_card.type--smiledelivery > div > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_smiledelivery > a > img')['alt']

    item_price = v.select_one('strong.text--price_seller').get_text()

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")


#vinsa
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'http://www.auction.co.kr/??pid=867&redirect=1'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="txtKeyword"]')
search_word = 'vinsa'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="core_header"]/div/div[1]/form/div[1]/input[2]').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')

list = soup.select('div.section--itemcard_info')

wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")
ws = wb.active

     
for v in list:        
    item_word = soup.head.find("meta", {"name" : "keywords"}).get('content')

    item_name = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_major > div.area--itemcard_title > span > a > span.text--title').text     

    try:   
        item_pq = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_score > ul > li > span.text--buycnt').text
    except Exception as e:
        continue
    
    try:
        item_seller = v.select_one('a.link--shop').get_text()  
        new_item_seller = item_seller.replace('판매자', '')
    except Exception as e:
        new_item_seller = v.select_one('#section--inner_content_body_container > div > div.component.component--item_card.type--smiledelivery > div > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_smiledelivery > a > img')['alt']


    item_price = v.select_one('strong.text--price_seller').get_text()

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")


#보스토 T1060
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'http://www.auction.co.kr/??pid=867&redirect=1'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="txtKeyword"]')
search_word = '보스토 T1060'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="core_header"]/div/div[1]/form/div[1]/input[2]').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')

list = soup.select('div.section--itemcard_info')

wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")
ws = wb.active

     
for v in list:        
    item_word = soup.head.find("meta", {"name" : "keywords"}).get('content')

    item_name = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_major > div.area--itemcard_title > span > a > span.text--title').text     

    try:   
        item_pq = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_score > ul > li > span.text--buycnt').text
    except Exception as e:
        continue
    
    try:
        item_seller = v.select_one('a.link--shop').get_text()  
        new_item_seller = item_seller.replace('판매자', '')
    except Exception as e:
        new_item_seller = v.select_one('#section--inner_content_body_container > div > div.component.component--item_card.type--smiledelivery > div > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_smiledelivery > a > img')['alt']


    item_price = v.select_one('strong.text--price_seller').get_text()

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")


#베이크 A15
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'http://www.auction.co.kr/??pid=867&redirect=1'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="txtKeyword"]')
search_word = '베이크 A15'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="core_header"]/div/div[1]/form/div[1]/input[2]').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')

list = soup.select('div.section--itemcard_info')

wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")
ws = wb.active

     
for v in list:        
    item_word = soup.head.find("meta", {"name" : "keywords"}).get('content')

    item_name = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_major > div.area--itemcard_title > span > a > span.text--title').text     

    try:   
        item_pq = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_score > ul > li > span.text--buycnt').text
    except Exception as e:
        continue
    
    try:
        item_seller = v.select_one('a.link--shop').get_text()  
        new_item_seller = item_seller.replace('판매자', '')
    except Exception as e:
        new_item_seller = v.select_one('#section--inner_content_body_container > div > div.component.component--item_card.type--smiledelivery > div > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_smiledelivery > a > img')['alt']


    item_price = v.select_one('strong.text--price_seller').get_text()

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")


#베이크 A30
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'http://www.auction.co.kr/??pid=867&redirect=1'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="txtKeyword"]')
search_word = '베이크 A30'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="core_header"]/div/div[1]/form/div[1]/input[2]').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')

list = soup.select('div.section--itemcard_info')

wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")
ws = wb.active

     
for v in list:        
    item_word = soup.head.find("meta", {"name" : "keywords"}).get('content')

    item_name = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_major > div.area--itemcard_title > span > a > span.text--title').text     

    try:   
        item_pq = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_score > ul > li > span.text--buycnt').text
    except Exception as e:
        continue
    
    try:
        item_seller = v.select_one('a.link--shop').get_text()  
        new_item_seller = item_seller.replace('판매자', '')
    except Exception as e:
        new_item_seller = v.select_one('#section--inner_content_body_container > div > div.component.component--item_card.type--smiledelivery > div > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_smiledelivery > a > img')['alt']


    item_price = v.select_one('strong.text--price_seller').get_text()

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")


#베이크 A50
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'http://www.auction.co.kr/??pid=867&redirect=1'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="txtKeyword"]')
search_word = '베이크 A50'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="core_header"]/div/div[1]/form/div[1]/input[2]').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')

list = soup.select('div.section--itemcard_info')

wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")
ws = wb.active

     
for v in list:        
    item_word = soup.head.find("meta", {"name" : "keywords"}).get('content')

    item_name = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_major > div.area--itemcard_title > span > a > span.text--title').text     

    try:   
        item_pq = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_score > ul > li > span.text--buycnt').text
    except Exception as e:
        continue
    
    try:
        item_seller = v.select_one('a.link--shop').get_text()  
        new_item_seller = item_seller.replace('판매자', '')
    except Exception as e:
        new_item_seller = v.select_one('#section--inner_content_body_container > div > div.component.component--item_card.type--smiledelivery > div > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_smiledelivery > a > img')['alt']


    item_price = v.select_one('strong.text--price_seller').get_text()

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")


#베이크 VK640
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'http://www.auction.co.kr/??pid=867&redirect=1'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="txtKeyword"]')
search_word = '베이크 VK640'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="core_header"]/div/div[1]/form/div[1]/input[2]').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')

list = soup.select('div.section--itemcard_info')

wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")
ws = wb.active

     
for v in list:        
    item_word = soup.head.find("meta", {"name" : "keywords"}).get('content')

    item_name = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_major > div.area--itemcard_title > span > a > span.text--title').text     

    try:   
        item_pq = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_score > ul > li > span.text--buycnt').text
    except Exception as e:
        continue
    
    try:
        item_seller = v.select_one('a.link--shop').get_text()  
        new_item_seller = item_seller.replace('판매자', '')
    except Exception as e:
        new_item_seller = v.select_one('#section--inner_content_body_container > div > div.component.component--item_card.type--smiledelivery > div > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_smiledelivery > a > img')['alt']


    item_price = v.select_one('strong.text--price_seller').get_text()

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")


#베이크 VK1060
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'http://www.auction.co.kr/??pid=867&redirect=1'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="txtKeyword"]')
search_word = '베이크 VK1060'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="core_header"]/div/div[1]/form/div[1]/input[2]').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')

list = soup.select('div.section--itemcard_info')

wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")
ws = wb.active

     
for v in list:        
    item_word = soup.head.find("meta", {"name" : "keywords"}).get('content')

    item_name = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_major > div.area--itemcard_title > span > a > span.text--title').text     

    try:   
        item_pq = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_score > ul > li > span.text--buycnt').text
    except Exception as e:
        continue
    
    try:
        item_seller = v.select_one('a.link--shop').get_text()  
        new_item_seller = item_seller.replace('판매자', '')
    except Exception as e:
        new_item_seller = v.select_one('#section--inner_content_body_container > div > div.component.component--item_card.type--smiledelivery > div > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_smiledelivery > a > img')['alt']


    item_price = v.select_one('strong.text--price_seller').get_text()

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")


#베이크 S640
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'http://www.auction.co.kr/??pid=867&redirect=1'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="txtKeyword"]')
search_word = '베이크 S640'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="core_header"]/div/div[1]/form/div[1]/input[2]').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')

list = soup.select('div.section--itemcard_info')

wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")
ws = wb.active

     
for v in list:        
    item_word = soup.head.find("meta", {"name" : "keywords"}).get('content')

    item_name = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_major > div.area--itemcard_title > span > a > span.text--title').text     

    try:   
        item_pq = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_score > ul > li > span.text--buycnt').text
    except Exception as e:
        continue
    
    try:
        item_seller = v.select_one('a.link--shop').get_text()  
        new_item_seller = item_seller.replace('판매자', '')
    except Exception as e:
        new_item_seller = v.select_one('#section--inner_content_body_container > div > div.component.component--item_card.type--smiledelivery > div > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_smiledelivery > a > img')['alt']


    item_price = v.select_one('strong.text--price_seller').get_text()

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")


#이지드로잉 1060plus
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'http://www.auction.co.kr/??pid=867&redirect=1'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="txtKeyword"]')
search_word = '이지드로잉 1060plus'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="core_header"]/div/div[1]/form/div[1]/input[2]').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')

list = soup.select('div.section--itemcard_info')

wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")
ws = wb.active

     
for v in list:        
    item_word = soup.head.find("meta", {"name" : "keywords"}).get('content')

    item_name = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_major > div.area--itemcard_title > span > a > span.text--title').text     

    try:   
        item_pq = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_score > ul > li > span.text--buycnt').text
    except Exception as e:
        continue
    
    try:
        item_seller = v.select_one('a.link--shop').get_text()  
        new_item_seller = item_seller.replace('판매자', '')
    except Exception as e:
        new_item_seller = v.select_one('#section--inner_content_body_container > div > div.component.component--item_card.type--smiledelivery > div > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_smiledelivery > a > img')['alt']


    item_price = v.select_one('strong.text--price_seller').get_text()

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")


#이지드로잉 노트
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'http://www.auction.co.kr/??pid=867&redirect=1'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="txtKeyword"]')
search_word = '이지드로잉 노트'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="core_header"]/div/div[1]/form/div[1]/input[2]').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')

list = soup.select('div.section--itemcard_info')

wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")
ws = wb.active

     
for v in list:        
    item_word = soup.head.find("meta", {"name" : "keywords"}).get('content')

    item_name = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_major > div.area--itemcard_title > span > a > span.text--title').text     

    try:   
        item_pq = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_score > ul > li > span.text--buycnt').text
    except Exception as e:
        continue
    
    try:
        item_seller = v.select_one('a.link--shop').get_text()  
        new_item_seller = item_seller.replace('판매자', '')
    except Exception as e:
        new_item_seller = v.select_one('#section--inner_content_body_container > div > div.component.component--item_card.type--smiledelivery > div > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_smiledelivery > a > img')['alt']


    item_price = v.select_one('strong.text--price_seller').get_text()

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")


#이지드로잉 지니
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'http://www.auction.co.kr/??pid=867&redirect=1'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="txtKeyword"]')
search_word = '이지드로잉 지니'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="core_header"]/div/div[1]/form/div[1]/input[2]').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')

list = soup.select('div.section--itemcard_info')

wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")
ws = wb.active

     
for v in list:        
    item_word = soup.head.find("meta", {"name" : "keywords"}).get('content')

    item_name = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_major > div.area--itemcard_title > span > a > span.text--title').text     

    try:   
        item_pq = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_score > ul > li > span.text--buycnt').text
    except Exception as e:
        continue
    
    try:
        item_seller = v.select_one('a.link--shop').get_text()  
        new_item_seller = item_seller.replace('판매자', '')
    except Exception as e:
        new_item_seller = v.select_one('#section--inner_content_body_container > div > div.component.component--item_card.type--smiledelivery > div > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_smiledelivery > a > img')['alt']


    item_price = v.select_one('strong.text--price_seller').get_text()

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")


#XPPEN DECO01
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'http://www.auction.co.kr/??pid=867&redirect=1'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="txtKeyword"]')
search_word = 'XPPEN DECO01'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="core_header"]/div/div[1]/form/div[1]/input[2]').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')

list = soup.select('div.section--itemcard_info')

wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")
ws = wb.active

     
for v in list:        
    item_word = soup.head.find("meta", {"name" : "keywords"}).get('content')

    item_name = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_major > div.area--itemcard_title > span > a > span.text--title').text     

    try:   
        item_pq = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_score > ul > li > span.text--buycnt').text
    except Exception as e:
        continue
    
    try:
        item_seller = v.select_one('a.link--shop').get_text()  
        new_item_seller = item_seller.replace('판매자', '')
    except Exception as e:
        new_item_seller = v.select_one('#section--inner_content_body_container > div > div.component.component--item_card.type--smiledelivery > div > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_smiledelivery > a > img')['alt']


    item_price = v.select_one('strong.text--price_seller').get_text()

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")


#XPPEN DECO02
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'http://www.auction.co.kr/??pid=867&redirect=1'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="txtKeyword"]')
search_word = 'XPPEN DECO02'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="core_header"]/div/div[1]/form/div[1]/input[2]').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')

list = soup.select('div.section--itemcard_info')

wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")
ws = wb.active

     
for v in list:        
    item_word = soup.head.find("meta", {"name" : "keywords"}).get('content')

    item_name = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_major > div.area--itemcard_title > span > a > span.text--title').text     

    try:   
        item_pq = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_score > ul > li > span.text--buycnt').text
    except Exception as e:
        continue
    
    try:
        item_seller = v.select_one('a.link--shop').get_text()  
        new_item_seller = item_seller.replace('판매자', '')
    except Exception as e:
        new_item_seller = v.select_one('#section--inner_content_body_container > div > div.component.component--item_card.type--smiledelivery > div > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_smiledelivery > a > img')['alt']


    item_price = v.select_one('strong.text--price_seller').get_text()

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")


#XPPEN Deco Pro S
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'http://www.auction.co.kr/??pid=867&redirect=1'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="txtKeyword"]')
search_word = 'XPPEN Deco Pro S'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="core_header"]/div/div[1]/form/div[1]/input[2]').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')

list = soup.select('div.section--itemcard_info')

wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")
ws = wb.active

     
for v in list:        
    item_word = soup.head.find("meta", {"name" : "keywords"}).get('content')
    new_item_word = item_word.replace('Deco Pro S', 'DecoProS')

    item_name = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_major > div.area--itemcard_title > span > a > span.text--title').text     

    try:   
        item_pq = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_score > ul > li > span.text--buycnt').text
    except Exception as e:
        continue

    try:
        item_seller = v.select_one('a.link--shop').get_text()  
        new_item_seller = item_seller.replace('판매자', '')
    except Exception as e:
        new_item_seller = v.select_one('#section--inner_content_body_container > div > div.component.component--item_card.type--smiledelivery > div > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_smiledelivery > a > img')['alt']

    item_price = v.select_one('strong.text--price_seller').get_text()

    print(new_item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([new_item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")


#XPPEN Deco Pro M
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'http://www.auction.co.kr/??pid=867&redirect=1'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="txtKeyword"]')
search_word = 'XPPEN Deco Pro M'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="core_header"]/div/div[1]/form/div[1]/input[2]').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')

list = soup.select('div.section--itemcard_info')

wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")
ws = wb.active

     
for v in list:        
    item_word = soup.head.find("meta", {"name" : "keywords"}).get('content')
    new_item_word = item_word.replace('Deco Pro M', 'DecoProM')

    item_name = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_major > div.area--itemcard_title > span > a > span.text--title').text     

    try:   
        item_pq = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_score > ul > li > span.text--buycnt').text
    except Exception as e:
        continue
    
    try:
        item_seller = v.select_one('a.link--shop').get_text()  
        new_item_seller = item_seller.replace('판매자', '')
    except Exception as e:
        new_item_seller = v.select_one('#section--inner_content_body_container > div > div.component.component--item_card.type--smiledelivery > div > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_smiledelivery > a > img')['alt']


    item_price = v.select_one('strong.text--price_seller').get_text()

    print(new_item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([new_item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")


#휴이온 RTS300
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'http://www.auction.co.kr/??pid=867&redirect=1'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="txtKeyword"]')
search_word = '휴이온 RTS300'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="core_header"]/div/div[1]/form/div[1]/input[2]').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')

list = soup.select('div.section--itemcard_info')

wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")
ws = wb.active

     
for v in list:        
    item_word = soup.head.find("meta", {"name" : "keywords"}).get('content')

    item_name = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_major > div.area--itemcard_title > span > a > span.text--title').text     

    try:   
        item_pq = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_score > ul > li > span.text--buycnt').text
    except Exception as e:
        continue
    
    try:
        item_seller = v.select_one('a.link--shop').get_text()  
        new_item_seller = item_seller.replace('판매자', '')
    except Exception as e:
        new_item_seller = v.select_one('#section--inner_content_body_container > div > div.component.component--item_card.type--smiledelivery > div > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_smiledelivery > a > img')['alt']


    item_price = v.select_one('strong.text--price_seller').get_text()

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")


#휴이온 RTM500
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'http://www.auction.co.kr/??pid=867&redirect=1'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="txtKeyword"]')
search_word = '휴이온 RTM500'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="core_header"]/div/div[1]/form/div[1]/input[2]').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')

list = soup.select('div.section--itemcard_info')

wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")
ws = wb.active

     
for v in list:        
    item_word = soup.head.find("meta", {"name" : "keywords"}).get('content')

    item_name = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_major > div.area--itemcard_title > span > a > span.text--title').text     

    try:   
        item_pq = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_score > ul > li > span.text--buycnt').text
    except Exception as e:
        continue
    
    try:
        item_seller = v.select_one('a.link--shop').get_text()  
        new_item_seller = item_seller.replace('판매자', '')
    except Exception as e:
        new_item_seller = v.select_one('#section--inner_content_body_container > div > div.component.component--item_card.type--smiledelivery > div > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_smiledelivery > a > img')['alt']


    item_price = v.select_one('strong.text--price_seller').get_text()

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")


#휴이온 RTP700
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'http://www.auction.co.kr/??pid=867&redirect=1'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="txtKeyword"]')
search_word = '휴이온 RTP700'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="core_header"]/div/div[1]/form/div[1]/input[2]').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')

list = soup.select('div.section--itemcard_info')

wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")
ws = wb.active

     
for v in list:        
    item_word = soup.head.find("meta", {"name" : "keywords"}).get('content')

    item_name = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_major > div.area--itemcard_title > span > a > span.text--title').text     

    try:   
        item_pq = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_score > ul > li > span.text--buycnt').text
    except Exception as e:
        continue
    
    try:
        item_seller = v.select_one('a.link--shop').get_text()  
        new_item_seller = item_seller.replace('판매자', '')
    except Exception as e:
        new_item_seller = v.select_one('#section--inner_content_body_container > div > div.component.component--item_card.type--smiledelivery > div > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_smiledelivery > a > img')['alt']


    item_price = v.select_one('strong.text--price_seller').get_text()

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")


#휴이온 H640P
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'http://www.auction.co.kr/??pid=867&redirect=1'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="txtKeyword"]')
search_word = '휴이온 H640P'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="core_header"]/div/div[1]/form/div[1]/input[2]').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')

list = soup.select('div.section--itemcard_info')

wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")
ws = wb.active

     
for v in list:        
    item_word = soup.head.find("meta", {"name" : "keywords"}).get('content')

    item_name = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_major > div.area--itemcard_title > span > a > span.text--title').text     

    try:   
        item_pq = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_score > ul > li > span.text--buycnt').text
    except Exception as e:
        continue
    
    try:
        item_seller = v.select_one('a.link--shop').get_text()  
        new_item_seller = item_seller.replace('판매자', '')
    except Exception as e:
        new_item_seller = v.select_one('#section--inner_content_body_container > div > div.component.component--item_card.type--smiledelivery > div > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_smiledelivery > a > img')['alt']


    item_price = v.select_one('strong.text--price_seller').get_text()

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")


#휴이온 H950P
chrome_options = Options()
driver = webdriver.Chrome('chromedriver.exe')
driver.maximize_window()

URL = 'http://www.auction.co.kr/??pid=867&redirect=1'


driver.get(URL)

search_box = driver.find_element(By.XPATH,'//*[@id="txtKeyword"]')
search_word = '휴이온 H950P'
search_box.send_keys(search_word)
driver.find_element(By.XPATH,'//*[@id="core_header"]/div/div[1]/form/div[1]/input[2]').click()

pre_scrollHeight = driver.execute_script("return document.body.scrollHeight") #이전 페이지 높이
interval = 2 #sleep time

while True:
    #스크롤을 가장 아래로 내림
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(interval)    #페이지 로딩 대기
    curr_scrollHeight = driver.execute_script("return document.body.scrollHeight")   #현재 높이

    #현재 높이 과거 높이 비교
    if pre_scrollHeight == curr_scrollHeight:
        break

    pre_scrollHeight = curr_scrollHeight
   

soup = BeautifulSoup(driver.page_source, 'html.parser')

list = soup.select('div.section--itemcard_info')

wb = op.load_workbook(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")
ws = wb.active

     
for v in list:        
    item_word = soup.head.find("meta", {"name" : "keywords"}).get('content')

    item_name = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_major > div.area--itemcard_title > span > a > span.text--title').text     

    try:   
        item_pq = v.select_one('#section--inner_content_body_container > div > div > div.itemcard > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_info_score > ul > li > span.text--buycnt').text
    except Exception as e:
        continue
    
    try:
        item_seller = v.select_one('a.link--shop').get_text()  
        new_item_seller = item_seller.replace('판매자', '')
    except Exception as e:
        new_item_seller = v.select_one('#section--inner_content_body_container > div > div.component.component--item_card.type--smiledelivery > div > div.section--itemcard > div.section--itemcard_info > div.section--itemcard_smiledelivery > a > img')['alt']


    item_price = v.select_one('strong.text--price_seller').get_text()

    print(item_word, item_name,item_pq,new_item_seller,item_price, sep=',')
    
    ws.append([item_word, item_name, item_pq, new_item_seller, item_price])
    wb.save(r"C:\Users\user\Desktop\PythonWorkSpace\month_auction.xlsx")